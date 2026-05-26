"""
Tesouraria INJ — Sistema de Gestão Financeira Multi-Congregação
Igreja do Nome de Jesus
"""
from flask import (Flask, render_template, request, redirect, url_for,
                   jsonify, flash, send_file, session, g)
import sqlite3, os, hashlib, secrets
from datetime import datetime
import json, io

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'data', 'tesouraria.db')

# ─────────────────────────────────────────────
#  DADOS FIXOS
# ─────────────────────────────────────────────
MESES = ['janeiro','fevereiro','março','abril','maio','junho',
         'julho','agosto','setembro','outubro','novembro','dezembro']

ANOS  = list(range(2020, 2035))

LIVROS = ['FundoLocal', 'Projetos', 'Envios', 'LiqDízimos']

LIVROS_LABELS = {
    'FundoLocal':  'Fundo Local',
    'Projetos':    'Projetos',
    'Envios':      'Envios / Missões',
    'LiqDízimos':  'Liquidação de Dízimos',
}

CATEGORIAS = {
    'FundoLocal': {
        'Entrada': sorted([
            'Comitê de Família','Contribuição do pastor para projetos',
            'Contribuição Irmão para pastor assembléia','Marcha missionária',
            'Of. Ação De Graças','Of. Cultos De Quintas','Of. Cultos De Terças',
            'Of. Cultos Especiais','Of. Cultos Sábados','Of. Domingos à Noite',
            'Of. Domingos E.B.D','Of. Na Refam','Of. Social (Casos Especiais)',
            'Oferta de Cavalheiros','Oferta de música','Oferta de obra social',
            'Outras Ofertas','Parceiros missionários','Voto do talento',
            'Votos','Votos Para Construção','%DízimosFundoLocal',
        ]),
        'Saída': sorted([
            'Aluguel de Casa Pastoral','Aluguel de Caçamba',
            'Aluguel de Maq. Construção','Aluguel Salão','Atividades legales',
            'AyudaPastorAux','Cafeteria/Restaurante','Completar uma Oferta',
            'Compra de Instrumentos','Compra de Materias de Construção',
            'Compra de Som','Compra de Utiles de limpeza',
            'Comprar artigos de papelaria','Comprar botijão de gás',
            'Construção E Reforma','DízimosDoAuxilioEmergencial','DízimosExtras',
            'Documentação E Escrituras','Evangelismo','Eventos',
            'Eventos E Alimentação','Manutenção Coz. Da Igreja',
            'Manutenção de Equipamento','Manutenção de Predio',
            'Manutenção de Som','Manutenção Do Prédio','Materiais P/ Escritório',
            'Obra Social','Of. Damas Miss. Fim Ano','Of. Damas Miss. Metade Ano',
            'Of. Dia Missionário','Of. Distrital','Of. Jovens Miss. Fim Ano',
            'Of. Jovens Miss. Metade Ano','Of. Raios De Luz',
            'Of. Safras Para Cristo','Otras Compras pra o Templo',
            'Otras Saídas de Projetos','Outros Fundo Local','Outros Envios',
            'Pago de Parcela','Pago Internet','Pago Mão de Obra',
            'produtos de limpeza','Projetos da diretoria local',
            'Serviços Públicos','Trabalho C/ Crianças','Trabalho C/ Damas',
            'Trabalho C/ Jovens','Transferência para o fundo de construção',
            'Transporte Urbano','Viáticos','Viaticos assambleia','%DízimoDosDízimos',
        ]),
    },
    'Projetos': {
        'Entrada': sorted([
            'Atividades - Cantina','Brechós','Eventos','Fundo de Transporte',
            'Outros Entradas de Projetos','Vendas','Votos','Votos Para Construção',
        ]),
        'Saída': sorted([
            'Aluguel de Caçamba','Aluguel de Maq. Construção',
            'Compra de Instrumentos','Compra de Materias de Construção',
            'Compra de Som','Compra de Utiles de limpeza',
            'DízimosDoAuxilioEmergencial','DízimosExtras',
            'Manutenção de Predio','Manutenção de Som',
            'Of. Damas Miss. Fim Ano','Of. Damas Miss. Metade Ano',
            'Of. Dia Missionário','Of. Distrital','Of. Jovens Miss. Fim Ano',
            'Of. Jovens Miss. Metade Ano','Of. Raios De Luz',
            'Of. Safras Para Cristo','Otras Compras pra o Templo',
            'Otras Saídas de Projetos','Outros Envios',
            'Pago de Parcela','Pago Mão de Obra','%DízimoDosDízimos',
        ]),
    },
    'Envios': {
        'Entrada': sorted([
            '%DízimoDosDízimos','Of. Dia Missionário','Of. Raios De Luz',
            'Of. Safras Para Cristo','Of. Jovens Miss. Metade Ano',
            'Of. Jovens Miss. Fim Ano','Of. Damas Miss. Metade Ano',
            'Of. Damas Miss. Fim Ano','Of. Distrital',
        ]),
        'Saída': sorted([
            '%DízimoDosDízimos','DízimosDoAuxilioEmergencial','DízimosExtras',
            'Of. Damas Miss. Fim Ano','Of. Damas Miss. Metade Ano',
            'Of. Dia Missionário','Of. Distrital','Of. Jovens Miss. Fim Ano',
            'Of. Jovens Miss. Metade Ano','Of. Raios De Luz',
            'Of. Safras Para Cristo','Outros Envios',
        ]),
    },
    'LiqDízimos': {
        'Entrada': sorted([
            'DízimoBruto','%DízimoDosDízimos','%DízimosFundoLocal',
            'LíquidoPastor','OtrosDízimos','DízimosDoAuxilioEmergencial','DízimosExtras',
        ]),
        'Saída': sorted([
            '%DízimoDosDízimos','%DízimosFundoLocal','LíquidoPastor',
            'Envío Nacional','Aporte ao Fundo Local','OtrosDízimos',
        ]),
    },
}

# ─────────────────────────────────────────────
#  BANCO DE DADOS
# ─────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def hash_senha(senha):
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    c = conn.cursor()

    # Tabela de congregações (cada uma é um tenant independente)
    c.execute('''CREATE TABLE IF NOT EXISTS congregacoes (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nome        TEXT    NOT NULL,
        usuario     TEXT    NOT NULL UNIQUE,
        senha_hash  TEXT    NOT NULL,
        is_admin    INTEGER DEFAULT 0,
        ativo       INTEGER DEFAULT 1,
        criado_em   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # Configurações por congregação
    c.execute('''CREATE TABLE IF NOT EXISTS config (
        cong_id INTEGER NOT NULL,
        key     TEXT    NOT NULL,
        value   TEXT,
        PRIMARY KEY (cong_id, key),
        FOREIGN KEY (cong_id) REFERENCES congregacoes(id)
    )''')

    # Transações
    c.execute('''CREATE TABLE IF NOT EXISTS transacoes (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        cong_id       INTEGER NOT NULL,
        livro         TEXT    NOT NULL,
        tipo          TEXT    NOT NULL,
        item          TEXT    NOT NULL,
        valor         REAL    NOT NULL,
        mes           TEXT    NOT NULL,
        ano           INTEGER NOT NULL,
        data_registro TEXT,
        obs           TEXT,
        created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (cong_id) REFERENCES congregacoes(id)
    )''')

    # Orçamento
    c.execute('''CREATE TABLE IF NOT EXISTS orcamento (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        cong_id    INTEGER NOT NULL,
        item       TEXT    NOT NULL,
        tipo       TEXT    NOT NULL,
        orcam_ano  REAL    DEFAULT 0,
        orcam_mes  REAL    DEFAULT 0,
        observacao TEXT,
        ano        INTEGER NOT NULL,
        UNIQUE(cong_id, item, tipo, ano),
        FOREIGN KEY (cong_id) REFERENCES congregacoes(id)
    )''')

    # Dízimos
    c.execute('''CREATE TABLE IF NOT EXISTS dizimos (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        cong_id           INTEGER NOT NULL,
        dizimo_bruto      REAL    NOT NULL,
        pct_nacional      REAL    DEFAULT 0.15,
        pct_fundo_local   REAL    DEFAULT 0.03,
        valor_nacional    REAL,
        valor_fundo_local REAL,
        valor_liquido     REAL,
        mes               TEXT    NOT NULL,
        ano               INTEGER NOT NULL,
        data_registro     TEXT,
        obs               TEXT,
        created_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (cong_id) REFERENCES congregacoes(id)
    )''')

    # Criar admin padrão se não existir
    admin = c.execute("SELECT id FROM congregacoes WHERE usuario='admin'").fetchone()
    if not admin:
        c.execute('''INSERT INTO congregacoes (nome, usuario, senha_hash, is_admin)
                     VALUES (?,?,?,1)''',
                  ('Administrador INJ', 'admin', hash_senha('admin123')))
        admin_id = c.lastrowid
        # Config padrão do admin
        for k, v in _config_defaults().items():
            c.execute('INSERT OR IGNORE INTO config (cong_id, key, value) VALUES (?,?,?)',
                      (admin_id, k, v))

    conn.commit()
    conn.close()

def _config_defaults(nome_cong=''):
    return {
        'nome_congregacao':   nome_cong or 'Igreja do Nome de Jesus',
        'nome_pastor':        '',
        'nome_tesoureiro':    '',
        'nome_auxiliar':      '',
        'ano_atual':          str(datetime.now().year),
        'saldo_ano_anterior': '0',
        'pct_nacional':       '0.15',
        'pct_fundo_local':    '0.03',
    }

# ── Helpers de sessão ────────────────────────
def get_cong_id():
    return session.get('cong_id')

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('cong_id'):
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'):
            flash('Acesso restrito ao administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def get_cfg():
    """Retorna config da congregação logada."""
    cid = get_cong_id()
    if not cid:
        return {}
    conn = get_db()
    rows = conn.execute('SELECT key, value FROM config WHERE cong_id=?', (cid,)).fetchall()
    conn.close()
    return {r['key']: r['value'] for r in rows}

def save_cfg(data: dict):
    cid = get_cong_id()
    conn = get_db()
    for k, v in data.items():
        conn.execute('INSERT OR REPLACE INTO config (cong_id, key, value) VALUES (?,?,?)', (cid, k, v))
    conn.commit()
    conn.close()

# ─────────────────────────────────────────────
#  FILTROS JINJA
# ─────────────────────────────────────────────
def fmt_brl(value):
    try:
        return f"R$ {float(value):,.2f}".replace(',','X').replace('.',',').replace('X','.')
    except Exception:
        return "R$ 0,00"

app.jinja_env.filters['brl'] = fmt_brl
app.jinja_env.globals['now'] = datetime.now

# ─────────────────────────────────────────────
#  AUTH — LOGIN / LOGOUT
# ─────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('cong_id'):
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        usuario = request.form['usuario'].strip()
        senha   = request.form['senha']
        conn    = get_db()
        cong = conn.execute(
            'SELECT * FROM congregacoes WHERE usuario=? AND ativo=1', (usuario,)
        ).fetchone()
        conn.close()

        if cong and cong['senha_hash'] == hash_senha(senha):
            session.clear()
            session['cong_id']    = cong['id']
            session['cong_nome']  = cong['nome']
            session['usuario']    = cong['usuario']
            session['is_admin']   = bool(cong['is_admin'])
            next_url = request.args.get('next') or url_for('dashboard')
            return redirect(next_url)
        flash('Usuário ou senha incorretos.', 'danger')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Sessão encerrada.', 'info')
    return redirect(url_for('login'))

# ─────────────────────────────────────────────
#  ADMIN — GERENCIAR CONGREGAÇÕES
# ─────────────────────────────────────────────
@app.route('/admin')
@login_required
@admin_required
def admin():
    conn = get_db()
    congs = conn.execute('SELECT * FROM congregacoes ORDER BY nome').fetchall()
    conn.close()
    return render_template('admin.html', congs=congs, cfg=get_cfg())

@app.route('/admin/nova', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_nova_cong():
    if request.method == 'POST':
        nome    = request.form['nome'].strip()
        usuario = request.form['usuario'].strip().lower()
        senha   = request.form['senha']
        conn    = get_db()
        try:
            c = conn.cursor()
            c.execute('INSERT INTO congregacoes (nome, usuario, senha_hash) VALUES (?,?,?)',
                      (nome, usuario, hash_senha(senha)))
            cid = c.lastrowid
            for k, v in _config_defaults(nome).items():
                c.execute('INSERT OR IGNORE INTO config (cong_id, key, value) VALUES (?,?,?)',
                          (cid, k, v))
            conn.commit()
            flash(f'Congregação "{nome}" criada! Usuário: {usuario}', 'success')
        except sqlite3.IntegrityError:
            flash(f'Usuário "{usuario}" já existe. Escolha outro.', 'danger')
        finally:
            conn.close()
        return redirect(url_for('admin'))

    return render_template('admin_nova.html', cfg=get_cfg())

@app.route('/admin/<int:cid>/senha', methods=['POST'])
@login_required
@admin_required
def admin_reset_senha(cid):
    nova = request.form['nova_senha']
    conn = get_db()
    conn.execute('UPDATE congregacoes SET senha_hash=? WHERE id=?', (hash_senha(nova), cid))
    conn.commit()
    conn.close()
    flash('Senha atualizada!', 'success')
    return redirect(url_for('admin'))

@app.route('/admin/<int:cid>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_cong(cid):
    conn = get_db()
    conn.execute('UPDATE congregacoes SET ativo = 1 - ativo WHERE id=?', (cid,))
    conn.commit()
    conn.close()
    flash('Status atualizado!', 'success')
    return redirect(url_for('admin'))

# ─────────────────────────────────────────────
#  DASHBOARD
# ─────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    cfg = get_cfg()
    cid = get_cong_id()
    ano = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))
    conn = get_db()

    totais = conn.execute('''
        SELECT tipo, COALESCE(SUM(valor),0) as total
        FROM transacoes WHERE cong_id=? AND ano=? GROUP BY tipo
    ''', (cid, ano)).fetchall()
    total_e  = sum(r['total'] for r in totais if r['tipo'] == 'Entrada')
    total_s  = sum(r['total'] for r in totais if r['tipo'] == 'Saída')
    saldo_ant = float(cfg.get('saldo_ano_anterior', 0))
    saldo    = saldo_ant + total_e - total_s

    por_livro = conn.execute('''
        SELECT livro,
               SUM(CASE WHEN tipo='Entrada' THEN valor ELSE 0 END) as entradas,
               SUM(CASE WHEN tipo='Saída'   THEN valor ELSE 0 END) as saidas
        FROM transacoes WHERE cong_id=? AND ano=? GROUP BY livro
    ''', (cid, ano)).fetchall()

    fluxo_mensal = []
    for mes in MESES:
        row = conn.execute('''
            SELECT SUM(CASE WHEN tipo='Entrada' THEN valor ELSE 0 END) as e,
                   SUM(CASE WHEN tipo='Saída'   THEN valor ELSE 0 END) as s
            FROM transacoes WHERE cong_id=? AND ano=? AND mes=?
        ''', (cid, ano, mes)).fetchone()
        fluxo_mensal.append({'mes': mes[:3].capitalize(), 'e': round(row['e'] or 0, 2), 's': round(row['s'] or 0, 2)})

    top_saidas = conn.execute('''
        SELECT item, SUM(valor) as total FROM transacoes
        WHERE cong_id=? AND ano=? AND tipo='Saída'
        GROUP BY item ORDER BY total DESC LIMIT 7
    ''', (cid, ano)).fetchall()

    recentes = conn.execute('''
        SELECT * FROM transacoes WHERE cong_id=? AND ano=? ORDER BY id DESC LIMIT 10
    ''', (cid, ano)).fetchall()

    conn.close()
    return render_template('dashboard.html',
        cfg=cfg, ano=ano, anos=ANOS,
        total_e=total_e, total_s=total_s, saldo=saldo, saldo_ant=saldo_ant,
        por_livro=por_livro,
        fluxo_mensal=json.dumps(fluxo_mensal),
        top_saidas=json.dumps([{'item': r['item'], 'total': round(r['total'],2)} for r in top_saidas]),
        recentes=recentes, livros_labels=LIVROS_LABELS,
    )

# ─────────────────────────────────────────────
#  LANÇAMENTOS
# ─────────────────────────────────────────────
@app.route('/lancamentos', methods=['GET', 'POST'])
@login_required
def lancamentos():
    cfg     = get_cfg()
    cid     = get_cong_id()
    ano_sel = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))
    mes_sel = request.args.get('mes', '')

    if request.method == 'POST':
        livro = request.form['livro']
        tipo  = request.form['tipo']
        item  = request.form['item']
        valor = request.form['valor'].replace(',', '.')
        mes   = request.form['mes']
        ano   = request.form['ano']
        data  = request.form.get('data_registro', datetime.now().strftime('%d/%m/%Y'))
        obs   = request.form.get('obs', '')
        conn  = get_db()
        conn.execute('''INSERT INTO transacoes (cong_id,livro,tipo,item,valor,mes,ano,data_registro,obs)
                        VALUES (?,?,?,?,?,?,?,?,?)''',
                     (cid, livro, tipo, item, float(valor), mes, int(ano), data, obs))
        conn.commit(); conn.close()
        flash(f'Lançamento de {fmt_brl(valor)} registrado!', 'success')
        return redirect(url_for('lancamentos', ano=ano, mes=mes))

    conn = get_db()
    q = 'SELECT * FROM transacoes WHERE cong_id=? AND ano=?'
    p = [cid, ano_sel]
    if mes_sel:
        q += ' AND mes=?'; p.append(mes_sel)
    q += ' ORDER BY id DESC'
    lista = conn.execute(q, p).fetchall()
    total_e = sum(t['valor'] for t in lista if t['tipo'] == 'Entrada')
    total_s = sum(t['valor'] for t in lista if t['tipo'] == 'Saída')
    conn.close()

    return render_template('lancamentos.html',
        cfg=cfg, categorias=CATEGORIAS, meses=MESES, anos=ANOS, livros=LIVROS,
        livros_labels=LIVROS_LABELS, transacoes=lista,
        ano_sel=ano_sel, mes_sel=mes_sel, total_e=total_e, total_s=total_s,
    )

@app.route('/lancamentos/<int:tid>/editar', methods=['GET', 'POST'])
@login_required
def editar_lancamento(tid):
    cfg = get_cfg(); cid = get_cong_id()
    conn = get_db()
    t = conn.execute('SELECT * FROM transacoes WHERE id=? AND cong_id=?', (tid, cid)).fetchone()
    if not t:
        conn.close(); flash('Não encontrado.', 'danger')
        return redirect(url_for('lancamentos'))

    if request.method == 'POST':
        conn.execute('''UPDATE transacoes SET livro=?,tipo=?,item=?,valor=?,mes=?,ano=?,data_registro=?,obs=?
                        WHERE id=? AND cong_id=?''',
                     (request.form['livro'], request.form['tipo'], request.form['item'],
                      float(request.form['valor'].replace(',','.')),
                      request.form['mes'], int(request.form['ano']),
                      request.form.get('data_registro',''), request.form.get('obs',''),
                      tid, cid))
        conn.commit(); conn.close()
        flash('Lançamento atualizado!', 'success')
        return redirect(url_for('lancamentos', ano=request.form['ano']))
    conn.close()
    return render_template('editar_lancamento.html',
        cfg=cfg, t=t, categorias=CATEGORIAS, meses=MESES, anos=ANOS,
        livros=LIVROS, livros_labels=LIVROS_LABELS)

@app.route('/lancamentos/<int:tid>/deletar', methods=['POST'])
@login_required
def deletar_lancamento(tid):
    cid = get_cong_id(); conn = get_db()
    t = conn.execute('SELECT ano FROM transacoes WHERE id=? AND cong_id=?', (tid, cid)).fetchone()
    ano = t['ano'] if t else datetime.now().year
    conn.execute('DELETE FROM transacoes WHERE id=? AND cong_id=?', (tid, cid))
    conn.commit(); conn.close()
    flash('Lançamento excluído.', 'warning')
    return redirect(url_for('lancamentos', ano=ano))

# ─────────────────────────────────────────────
#  DÍZIMOS
# ─────────────────────────────────────────────
@app.route('/dizimos', methods=['GET', 'POST'])
@login_required
def dizimos():
    cfg     = get_cfg(); cid = get_cong_id()
    pct_nac = float(cfg.get('pct_nacional', 0.15))
    pct_fl  = float(cfg.get('pct_fundo_local', 0.03))
    calculo = None
    ano_sel = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))

    if request.method == 'POST':
        bruto    = float(request.form['dizimo_bruto'].replace(',', '.'))
        pct_nac_f = float(request.form.get('pct_nacional', pct_nac))
        pct_fl_f  = float(request.form.get('pct_fundo_local', pct_fl))
        mes  = request.form['mes']; ano = int(request.form['ano'])
        obs  = request.form.get('obs', '')
        data = request.form.get('data_registro', datetime.now().strftime('%d/%m/%Y'))
        v_nac = round(bruto * pct_nac_f, 2)
        v_fl  = round(bruto * pct_fl_f, 2)
        v_liq = round(bruto - v_nac - v_fl, 2)
        conn  = get_db()
        conn.execute('''INSERT INTO dizimos
            (cong_id,dizimo_bruto,pct_nacional,pct_fundo_local,valor_nacional,valor_fundo_local,valor_liquido,mes,ano,data_registro,obs)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (cid, bruto, pct_nac_f, pct_fl_f, v_nac, v_fl, v_liq, mes, ano, data, obs))
        conn.commit(); conn.close()
        calculo = {'bruto': bruto, 'pct_nac': pct_nac_f, 'pct_fl': pct_fl_f,
                   'v_nac': v_nac, 'v_fl': v_fl, 'v_liquido': v_liq, 'mes': mes, 'ano': ano}
        flash('Dízimo registrado!', 'success')

    conn = get_db()
    historico = conn.execute('SELECT * FROM dizimos WHERE cong_id=? AND ano=? ORDER BY id DESC', (cid, ano_sel)).fetchall()
    conn.close()
    return render_template('dizimos.html',
        cfg=cfg, meses=MESES, anos=ANOS, calculo=calculo,
        pct_nac=pct_nac, pct_fl=pct_fl, historico=historico, ano_sel=ano_sel)

@app.route('/dizimos/<int:did>/deletar', methods=['POST'])
@login_required
def deletar_dizimo(did):
    cid = get_cong_id(); conn = get_db()
    conn.execute('DELETE FROM dizimos WHERE id=? AND cong_id=?', (did, cid))
    conn.commit(); conn.close()
    flash('Registro excluído.', 'warning')
    return redirect(url_for('dizimos'))

# ─────────────────────────────────────────────
#  ORÇAMENTO
# ─────────────────────────────────────────────
@app.route('/orcamento', methods=['GET', 'POST'])
@login_required
def orcamento():
    cfg     = get_cfg(); cid = get_cong_id()
    ano_sel = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))

    if request.method == 'POST':
        conn = get_db()
        items     = request.form.getlist('item[]')
        tipos     = request.form.getlist('tipo[]')
        anos_f    = request.form.getlist('ano[]')
        orcam_anos = request.form.getlist('orcam_ano[]')
        orcam_meses = request.form.getlist('orcam_mes[]')
        obs_list  = request.form.getlist('obs[]')
        for i in range(len(items)):
            if items[i]:
                conn.execute('''INSERT INTO orcamento (cong_id,item,tipo,orcam_ano,orcam_mes,observacao,ano)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(cong_id,item,tipo,ano) DO UPDATE SET
                        orcam_ano=excluded.orcam_ano, orcam_mes=excluded.orcam_mes, observacao=excluded.observacao''',
                    (cid, items[i], tipos[i], float(orcam_anos[i] or 0),
                     float(orcam_meses[i] or 0),
                     obs_list[i] if i < len(obs_list) else '',
                     int(anos_f[i]) if i < len(anos_f) else ano_sel))
        conn.commit(); conn.close()
        flash('Orçamento salvo!', 'success')
        return redirect(url_for('orcamento', ano=ano_sel))

    conn = get_db()
    orc_rows = conn.execute('SELECT * FROM orcamento WHERE cong_id=? AND ano=?', (cid, ano_sel)).fetchall()
    orc_map  = {(r['item'], r['tipo']): r for r in orc_rows}

    todos_e = sorted(set(i for l in CATEGORIAS.values() for i in l.get('Entrada', [])))
    todos_s = sorted(set(i for l in CATEGORIAS.values() for i in l.get('Saída', [])))

    def comparativo(itens, tipo):
        result = []
        for item in itens:
            real = conn.execute('''SELECT COALESCE(SUM(valor),0) as t FROM transacoes
                WHERE cong_id=? AND item=? AND tipo=? AND ano=?''', (cid, item, tipo, ano_sel)).fetchone()['t']
            orc  = orc_map.get((item, tipo))
            orcado = orc['orcam_ano'] if orc else 0
            if real > 0 or orcado > 0:
                pct = round(real / orcado * 100, 1) if orcado > 0 else 0
                result.append({'item': item, 'orcado': orcado, 'real': real, 'pct': pct})
        return result

    comp_e = comparativo(todos_e, 'Entrada')
    comp_s = comparativo(todos_s, 'Saída')
    conn.close()

    return render_template('orcamento.html',
        cfg=cfg, ano_sel=ano_sel, anos=ANOS, categorias=CATEGORIAS, livros=LIVROS,
        comparativo_e=comp_e, comparativo_s=comp_s,
        orc_map={k: dict(v) for k, v in orc_map.items()})

# ─────────────────────────────────────────────
#  RELATÓRIOS
# ─────────────────────────────────────────────
@app.route('/relatorio')
@login_required
def relatorio():
    cfg       = get_cfg(); cid = get_cong_id()
    ano_sel   = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))
    livro_sel = request.args.get('livro', '')
    tipo_sel  = request.args.get('tipo', '')
    item_sel  = request.args.get('item', '')
    mes_sel   = request.args.get('mes', '')
    conn      = get_db()

    q = 'SELECT * FROM transacoes WHERE cong_id=? AND ano=?'
    p = [cid, ano_sel]
    if livro_sel: q += ' AND livro=?';           p.append(livro_sel)
    if tipo_sel:  q += ' AND tipo=?';            p.append(tipo_sel)
    if item_sel:  q += ' AND item LIKE ?';       p.append(f'%{item_sel}%')
    if mes_sel:   q += ' AND mes=?';             p.append(mes_sel)
    q += ' ORDER BY mes, id'
    lista = conn.execute(q, p).fetchall()

    total_e  = sum(t['valor'] for t in lista if t['tipo'] == 'Entrada')
    total_s  = sum(t['valor'] for t in lista if t['tipo'] == 'Saída')
    superavit = total_e - total_s
    pct = round(superavit / total_e * 100, 2) if total_e > 0 else 0

    resumo_mensal = []
    for mes in MESES:
        q2 = 'SELECT tipo, COALESCE(SUM(valor),0) as t FROM transacoes WHERE cong_id=? AND ano=? AND mes=?'
        p2 = [cid, ano_sel, mes]
        if livro_sel: q2 += ' AND livro=?'; p2.append(livro_sel)
        q2 += ' GROUP BY tipo'
        rows = conn.execute(q2, p2).fetchall()
        e = sum(r['t'] for r in rows if r['tipo'] == 'Entrada')
        s = sum(r['t'] for r in rows if r['tipo'] == 'Saída')
        if e > 0 or s > 0:
            resumo_mensal.append({'mes': mes, 'entradas': e, 'saidas': s, 'saldo': e-s})

    resumo_livro = conn.execute('''
        SELECT livro,
               SUM(CASE WHEN tipo='Entrada' THEN valor ELSE 0 END) as e,
               SUM(CASE WHEN tipo='Saída'   THEN valor ELSE 0 END) as s
        FROM transacoes WHERE cong_id=? AND ano=? GROUP BY livro
    ''', (cid, ano_sel)).fetchall()

    items_disp = conn.execute(
        'SELECT DISTINCT item FROM transacoes WHERE cong_id=? AND ano=? ORDER BY item', (cid, ano_sel)
    ).fetchall()
    conn.close()

    return render_template('relatorio.html',
        cfg=cfg, anos=ANOS, meses=MESES, livros=LIVROS, livros_labels=LIVROS_LABELS,
        ano_sel=ano_sel, livro_sel=livro_sel, tipo_sel=tipo_sel,
        item_sel=item_sel, mes_sel=mes_sel, transacoes=lista,
        total_e=total_e, total_s=total_s, superavit=superavit, pct=pct,
        resumo_mensal=resumo_mensal, resumo_livro=resumo_livro,
        items_disponiveis=items_disp)

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES
# ─────────────────────────────────────────────
@app.route('/configuracoes', methods=['GET', 'POST'])
@login_required
def configuracoes():
    if request.method == 'POST':
        save_cfg({k: request.form.get(k, '') for k in
                  ['nome_congregacao','nome_pastor','nome_tesoureiro','nome_auxiliar',
                   'ano_atual','saldo_ano_anterior','pct_nacional','pct_fundo_local']})
        # Atualizar também o nome na tabela de congregações
        nome = request.form.get('nome_congregacao', '')
        if nome:
            conn = get_db()
            conn.execute('UPDATE congregacoes SET nome=? WHERE id=?', (nome, get_cong_id()))
            conn.commit(); conn.close()
        flash('Configurações salvas!', 'success')
        return redirect(url_for('configuracoes'))

    return render_template('configuracoes.html', cfg=get_cfg(), anos=ANOS)

# ─────────────────────────────────────────────
#  MINHA CONTA (trocar senha)
# ─────────────────────────────────────────────
@app.route('/minha-conta', methods=['GET', 'POST'])
@login_required
def minha_conta():
    cfg = get_cfg()
    if request.method == 'POST':
        atual = request.form['senha_atual']
        nova  = request.form['nova_senha']
        conf  = request.form['confirmar_senha']
        cid   = get_cong_id()
        conn  = get_db()
        cong  = conn.execute('SELECT * FROM congregacoes WHERE id=?', (cid,)).fetchone()
        if cong['senha_hash'] != hash_senha(atual):
            flash('Senha atual incorreta.', 'danger')
        elif nova != conf:
            flash('As senhas não coincidem.', 'danger')
        elif len(nova) < 6:
            flash('A nova senha deve ter pelo menos 6 caracteres.', 'danger')
        else:
            conn.execute('UPDATE congregacoes SET senha_hash=? WHERE id=?', (hash_senha(nova), cid))
            conn.commit()
            flash('Senha alterada com sucesso!', 'success')
        conn.close()
    return render_template('minha_conta.html', cfg=cfg)

# ─────────────────────────────────────────────
#  API
# ─────────────────────────────────────────────
@app.route('/api/categorias/<livro>/<tipo>')
@login_required
def api_categorias(livro, tipo):
    return jsonify(CATEGORIAS.get(livro, {}).get(tipo, []))

# ─────────────────────────────────────────────
#  EXPORTAR EXCEL
# ─────────────────────────────────────────────
@app.route('/exportar/excel')
@login_required
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl não instalado.', 'danger')
        return redirect(url_for('relatorio'))

    cfg     = get_cfg(); cid = get_cong_id()
    ano_sel = int(request.args.get('ano', cfg.get('ano_atual', datetime.now().year)))
    conn    = get_db()
    lista   = conn.execute(
        'SELECT * FROM transacoes WHERE cong_id=? AND ano=? ORDER BY mes,livro,tipo,item',
        (cid, ano_sel)).fetchall()

    wb  = openpyxl.Workbook()
    ws  = wb.active; ws.title = 'Extrato'
    azul   = PatternFill('solid', fgColor='1A3A5C')
    branco = Font(color='FFFFFF', bold=True)
    verde  = PatternFill('solid', fgColor='D4EFDF')
    verm   = PatternFill('solid', fgColor='FADBD8')

    cols = ['#','Livro','Tipo','Item','Valor R$','Mês','Ano','Data','Obs']
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(1, ci, col)
        cell.font = branco; cell.fill = azul
        cell.alignment = Alignment(horizontal='center')

    for ri, t in enumerate(lista, 2):
        ws.cell(ri,1,t['id']); ws.cell(ri,2,t['livro']); ws.cell(ri,3,t['tipo'])
        ws.cell(ri,4,t['item'])
        c = ws.cell(ri,5,round(t['valor'],2)); c.number_format='R$ #,##0.00'
        ws.cell(ri,6,t['mes']); ws.cell(ri,7,t['ano'])
        ws.cell(ri,8,t['data_registro'] or ''); ws.cell(ri,9,t['obs'] or '')
        fill = verde if t['tipo']=='Entrada' else verm
        for ci in range(1,10): ws.cell(ri,ci).fill = fill

    for col, w in zip(['A','B','C','D','E','F','G','H','I'],
                      [6,15,10,35,15,12,8,13,25]):
        ws.column_dimensions[col].width = w

    # Resumo mensal
    ws2 = wb.create_sheet('Resumo Mensal')
    ws2.cell(1,1,f'Resumo Mensal {ano_sel} — {cfg.get("nome_congregacao","")}').font = Font(bold=True,size=14)
    ws2.merge_cells('A1:E1')
    cols2 = ['Mês','Entradas R$','Saídas R$','Saldo R$','% Superávit']
    for ci, c2 in enumerate(cols2,1):
        cell = ws2.cell(2,ci,c2); cell.font=branco; cell.fill=azul
        cell.alignment=Alignment(horizontal='center')
    te=ts=0
    for ri, mes in enumerate(MESES,3):
        row = conn.execute('''SELECT SUM(CASE WHEN tipo='Entrada' THEN valor ELSE 0 END) as e,
               SUM(CASE WHEN tipo='Saída' THEN valor ELSE 0 END) as s
               FROM transacoes WHERE cong_id=? AND ano=? AND mes=?''', (cid,ano_sel,mes)).fetchone()
        e,s = (row['e'] or 0),(row['s'] or 0)
        ws2.cell(ri,1,mes.capitalize())
        for ci,v in enumerate([e,s,e-s],2):
            c2=ws2.cell(ri,ci,round(v,2)); c2.number_format='R$ #,##0.00'
        ws2.cell(ri,5,f'{round((e-s)/e*100,1) if e>0 else 0}%')
        te+=e; ts+=s
    tr=len(MESES)+3
    ws2.cell(tr,1,'TOTAL').font=Font(bold=True)
    for ci,v in enumerate([te,ts,te-ts],2):
        c2=ws2.cell(tr,ci,round(v,2)); c2.number_format='R$ #,##0.00'; c2.font=Font(bold=True)
    for col in ['A','B','C','D','E']: ws2.column_dimensions[col].width=18
    conn.close()

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Tesouraria_{cfg.get("nome_congregacao","INJ")}_{ano_sel}.xlsx')

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    import socket
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = '127.0.0.1'
    port = int(os.environ.get('PORT', 8080))
    print("\n" + "="*60)
    print("  🏛️  TESOURARIA INJ — Sistema Multi-Congregação")
    print("="*60)
    print(f"  💻  Neste computador : http://127.0.0.1:{port}")
    print(f"  🌐  Rede local (Wi-Fi): http://{local_ip}:{port}")
    print(f"  👤  Admin padrão     : usuario=admin  senha=admin123")
    print("  ⛔  Para parar: Ctrl+C")
    print("="*60 + "\n")
    app.run(debug=False, host='0.0.0.0', port=port)
